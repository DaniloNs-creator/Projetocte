# ==============================================================================
# SISTEMA DE PROCESSAMENTO DE CT-e 2026 - HÄFELE BRASIL
# Versão: 3.1 - Suporte a Múltiplos ZIPs e Processamento Massivo
# ==============================================================================

import streamlit as st
from datetime import datetime
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from typing import Optional, Dict, Any, List, Tuple
import chardet
from io import BytesIO
import time
import xml.etree.ElementTree as ET
import os
import traceback
import numpy as np
import re
import tempfile
import logging
import gc
import io
from pathlib import Path
import zipfile
import shutil
import concurrent.futures
from threading import Lock, Semaphore
import base64
import psutil

# ==============================================================================
# CONFIGURAÇÃO AUTOMÁTICA DO SERVIDOR STREAMLIT
# ==============================================================================
def setup_streamlit_config():
    try:
        os.makedirs(".streamlit", exist_ok=True)
        config_path = os.path.join(".streamlit", "config.toml")
        with open(config_path, "w", encoding="utf-8") as f:
            f.write("[server]\nmaxUploadSize = 2000\nmaxMessageSize = 2000\n")
    except Exception:
        pass

setup_streamlit_config()

# ==============================================================================
# CONFIGURAÇÃO INICIAL
# ==============================================================================
st.set_page_config(
    page_title="Sistema de Processamento CT-e 2026",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

CTE_NAMESPACES = {'cte': 'http://www.portalfiscal.inf.br/cte'}

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ==============================================================================
# COMPAT HELPER
# ==============================================================================
def _w(stretch: bool = True):
    try:
        import inspect
        sig = inspect.signature(st.dataframe)
        if "width" in sig.parameters and "use_container_width" not in sig.parameters:
            return {"width": "stretch" if stretch else "content"}
        else:
            return {"use_container_width": stretch}
    except Exception:
        return {"use_container_width": stretch}

_WS = _w(True)
_WC = _w(False)

# ==============================================================================
# SESSION STATE
# ==============================================================================
_defaults = {
    'cte_data': [],
    'cte_df': None,
    'processing_log': [],
    'total_processed': 0,
    'total_errors': 0,
    'total_files': 0,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ==============================================================================
# HELPERS UI
# ==============================================================================
def show_loading_animation(message="Processando..."):
    with st.spinner(message):
        pb = st.progress(0)
        for i in range(100):
            time.sleep(0.01)
            pb.progress(i + 1)
        pb.empty()

def show_processing_animation(message="Analisando dados..."):
    ph_container = st.empty()
    with ph_container.container():
        _, c, _ = st.columns([1, 2, 1])
        with c:
            st.info(f"⏳ {message}")
            sp = st.empty()
            chars = ["⣾","⣽","⣻","⢿","⡿","⣟","⣯","⣷"]
            for i in range(20):
                sp.markdown(
                    f"<div style='text-align:center;font-size:24px'>{chars[i%8]}</div>",
                    unsafe_allow_html=True,
                )
                time.sleep(0.1)
    ph_container.empty()

def show_success_animation(message="Concluído!"):
    ph_container = st.empty()
    with ph_container.container():
        st.success(f"✅ {message}")
        time.sleep(1.2)
    ph_container.empty()

def ph(html: str):
    st.markdown(html, unsafe_allow_html=True)

def page_header(icon: str, title: str, sub: str):
    ph(f"""
    <div class="ph-hdr">
        <span class="ph-icon">{icon}</span>
        <div>
            <div class="ph-title">{title}</div>
            <div class="ph-sub">{sub}</div>
        </div>
    </div>""")

def section_title(text: str):
    ph(f'<div class="stitle">{text}</div>')

def empty_state(icon: str, title: str, sub: str = ""):
    ph(f"""
    <div class="empty">
        <div class="empty-icon">{icon}</div>
        <div class="empty-title">{title}</div>
        <div class="empty-sub">{sub}</div>
    </div>""")

def status_ok(text: str):
    ph(f'<div class="sbox sbox-ok">✅ {text}</div>')

def status_warn(text: str):
    ph(f'<div class="sbox sbox-warn">⚠️ {text}</div>')

def status_err(text: str):
    ph(f'<div class="sbox sbox-err">❌ {text}</div>')

# ==============================================================================
# CSS
# ==============================================================================
def load_css():
    ph("""<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600&display=swap');

    :root{
        --navy:#0A0F1E;
        --blue-dark:#0F2040;
        --blue:#1E3A8A;
        --blue-m:#2563EB;
        --blue-l:#3B82F6;
        --blue-xl:#60A5FA;
        --blue-bg:#EFF6FF;
        --blue-b:#BFDBFE;
        --green:#059669;
        --green-l:#10B981;
        --green-bg:#D1FAE5;
        --amber:#D97706;
        --amber-bg:#FEF3C7;
        --red:#DC2626;
        --red-bg:#FEE2E2;
        --bg:#F0F4FA;
        --surface:#FFFFFF;
        --surface2:#F8FAFC;
        --border:#E2E8F0;
        --border-strong:#CBD5E1;
        --muted:#64748B;
        --muted2:#94A3B8;
        --text:#0F172A;
        --r:10px;
        --r-lg:16px;
        --r-xl:24px;
        --r-2xl:32px;
        --sh0:0 1px 3px rgba(0,0,0,.06);
        --sh1:0 2px 8px rgba(0,0,0,.08),0 1px 3px rgba(0,0,0,.05);
        --sh2:0 8px 24px rgba(0,0,0,.10),0 2px 8px rgba(0,0,0,.06);
        --sh3:0 20px 60px rgba(0,0,0,.14),0 4px 16px rgba(0,0,0,.08);
        --sh-blue:0 8px 32px rgba(37,99,235,.20);
        --tr:all .2s cubic-bezier(.4,0,.2,1);
        --glow:0 0 0 3px rgba(59,130,246,.25);
    }

    html,body,[class*="css"]{
        font-family:'Inter','Segoe UI',system-ui,sans-serif;
        -webkit-font-smoothing:antialiased;
        color:var(--text);
    }
    ::-webkit-scrollbar{width:6px;height:6px}
    ::-webkit-scrollbar-track{background:var(--bg);border-radius:10px}
    ::-webkit-scrollbar-thumb{background:var(--border-strong);border-radius:10px}
    ::-webkit-scrollbar-thumb:hover{background:var(--muted2)}
    .block-container{padding-top:1rem!important;padding-bottom:2rem!important;max-width:1400px!important;}

    .hero{
        position:relative;
        background:linear-gradient(135deg,#050D1F 0%,#0F2040 35%,#1E3A8A 65%,#1D4ED8 100%);
        border-radius:var(--r-2xl);
        padding:2.8rem 3.5rem 2.4rem;
        margin-bottom:1.6rem;
        text-align:center;
        overflow:hidden;
        border:1px solid rgba(255,255,255,.06);
        box-shadow:var(--sh3),var(--sh-blue);
    }
    .hero::before{
        content:'';position:absolute;inset:0;
        background-image:
            linear-gradient(rgba(255,255,255,.03) 1px,transparent 1px),
            linear-gradient(90deg,rgba(255,255,255,.03) 1px,transparent 1px);
        background-size:48px 48px;pointer-events:none;
    }
    .hero::after{
        content:'';position:absolute;right:-80px;top:-80px;
        width:340px;height:340px;border-radius:50%;
        background:radial-gradient(circle,rgba(96,165,250,.15) 0%,transparent 65%);
        pointer-events:none;
    }
    .hero-glow-left{
        position:absolute;left:-100px;bottom:-80px;
        width:280px;height:280px;border-radius:50%;
        background:radial-gradient(circle,rgba(16,185,129,.10) 0%,transparent 65%);
        pointer-events:none;
    }
    .hero-logo{
        max-width:180px;margin-bottom:1rem;
        filter:drop-shadow(0 4px 18px rgba(0,0,0,.40));
        position:relative;z-index:1;
        transition:var(--tr);
    }
    .hero-logo:hover{transform:scale(1.03);}
    .hero-title{
        font-size:2.2rem;font-weight:900;color:#fff;
        margin:0 0 .4rem;letter-spacing:-.8px;
        line-height:1.12;position:relative;z-index:1;
        text-shadow:0 2px 12px rgba(0,0,0,.3);
    }
    .hero-sub{
        font-size:.95rem;color:rgba(255,255,255,.65);
        margin:0 0 1.4rem;position:relative;z-index:1;
        letter-spacing:.1px;
    }
    .hero-chips{
        display:flex;justify-content:center;
        gap:.5rem;flex-wrap:wrap;position:relative;z-index:1;
    }
    .chip{
        display:inline-flex;align-items:center;gap:.3rem;
        background:rgba(255,255,255,.10);
        border:1px solid rgba(255,255,255,.20);
        color:rgba(255,255,255,.90);border-radius:20px;
        padding:.22rem .8rem;font-size:.74rem;
        font-weight:600;letter-spacing:.3px;
        transition:var(--tr);backdrop-filter:blur(6px);
    }
    .chip:hover{background:rgba(255,255,255,.20);transform:translateY(-1px);}

    .ph-hdr{
        display:flex;align-items:center;gap:1rem;
        background:var(--surface);
        border:1px solid var(--border);
        border-left:4px solid var(--blue-l);
        border-radius:var(--r);padding:.9rem 1.4rem;
        margin-bottom:1.2rem;box-shadow:var(--sh0);
        transition:var(--tr);
    }
    .ph-hdr:hover{box-shadow:var(--sh1);border-left-color:var(--blue-m);}
    .ph-icon{font-size:2rem;flex-shrink:0;line-height:1;}
    .ph-title{font-size:1.3rem;font-weight:800;color:var(--blue);line-height:1.2;}
    .ph-sub{font-size:.8rem;color:var(--muted);margin-top:.15rem;}

    .stitle{
        display:flex;align-items:center;
        font-size:.88rem;font-weight:700;
        color:var(--blue);
        padding:.5rem 0 .5rem .85rem;
        border-left:3px solid var(--blue-l);
        margin:1.1rem 0 .7rem;
        background:linear-gradient(90deg,rgba(59,130,246,.07),transparent 80%);
        border-radius:0 var(--r) var(--r) 0;
        letter-spacing:.2px;
    }

    .card{
        background:var(--surface);
        border-radius:var(--r-lg);
        border:1px solid var(--border);
        box-shadow:var(--sh1);
        padding:1.3rem 1.5rem;
        margin-bottom:1rem;
        transition:var(--tr);
    }
    .card:hover{box-shadow:var(--sh2);border-color:var(--blue-b);}
    .card-accent{border-top:3px solid var(--blue-l);}

    .uzone{
        background:linear-gradient(135deg,var(--blue-bg),#DBEAFE88);
        border:2px dashed #93C5FD;
        border-radius:var(--r-lg);padding:1.1rem 1rem;
        text-align:center;margin-bottom:.5rem;
        transition:var(--tr);cursor:pointer;
    }
    .uzone:hover{border-color:var(--blue-l);background:linear-gradient(135deg,#DBEAFE,#EFF6FF);}
    .uzone-icon{font-size:1.7rem;line-height:1;margin-bottom:.3rem;}
    .uzone-title{font-weight:700;color:var(--blue);font-size:.9rem;margin-top:.2rem;}
    .uzone-sub{font-size:.75rem;color:var(--muted);margin-top:.15rem;}

    .sbox{
        padding:.7rem 1.1rem;border-radius:var(--r);
        font-size:.88rem;font-weight:500;margin:.4rem 0;
        display:flex;align-items:center;gap:.5rem;
    }
    .sbox-ok{
        background:var(--green-bg);color:#065F46;
        border:1px solid #A7F3D0;border-left:3px solid var(--green);
    }
    .sbox-warn{
        background:var(--amber-bg);color:#78350F;
        border:1px solid #FDE68A;border-left:3px solid var(--amber);
    }
    .sbox-err{
        background:var(--red-bg);color:#991B1B;
        border:1px solid #FECACA;border-left:3px solid var(--red);
    }

    .lbadge{
        display:inline-flex;align-items:center;gap:.35rem;
        background:var(--blue-m);color:#fff;
        border-radius:var(--r);padding:.3rem .85rem;
        font-size:.78rem;font-weight:700;
        margin-top:.5rem;box-shadow:var(--sh-blue);
        letter-spacing:.2px;
    }

    .flabel{
        font-size:.76rem;font-weight:600;color:var(--muted);
        text-transform:uppercase;letter-spacing:.6px;margin-bottom:.3rem;
    }

    .empty{
        text-align:center;padding:3.5rem 1.5rem;
        color:var(--muted);border:2px dashed var(--border);
        border-radius:var(--r-xl);background:var(--surface2);
    }
    .empty-icon{font-size:3rem;margin-bottom:.6rem;opacity:.5;}
    .empty-title{font-size:1rem;font-weight:700;color:var(--muted2);margin-bottom:.3rem;}
    .empty-sub{font-size:.82rem;color:#CBD5E1;}

    .stTabs [data-baseweb="tab-list"]{
        gap:3px;background:var(--bg);
        border-radius:var(--r-lg);
        padding:5px;border:1px solid var(--border);
    }
    .stTabs [data-baseweb="tab"]{
        border-radius:8px;font-weight:600;font-size:.85rem;
        padding:.42rem 1rem;transition:var(--tr);color:var(--muted);
        border:none;
    }
    .stTabs [data-baseweb="tab"]:hover{
        color:var(--blue-m);background:rgba(59,130,246,.08);
    }
    .stTabs [aria-selected="true"]{
        background:var(--surface)!important;
        color:var(--blue)!important;
        box-shadow:var(--sh1)!important;
    }

    .stButton>button{
        border-radius:var(--r)!important;font-weight:600!important;
        font-size:.86rem!important;letter-spacing:.1px;
        transition:var(--tr)!important;
    }
    .stButton>button:hover{
        transform:translateY(-1px)!important;
        box-shadow:var(--sh2)!important;
    }
    .stButton>button:active{
        transform:translateY(0)!important;
        box-shadow:var(--sh0)!important;
    }
    .stButton>button[kind="primary"]{
        background:linear-gradient(135deg,var(--blue-m),var(--blue))!important;
        box-shadow:var(--sh-blue)!important;border:none!important;
    }
    .stButton>button[kind="primary"]:hover{
        background:linear-gradient(135deg,#1D4ED8,var(--blue))!important;
    }

    .streamlit-expanderHeader{
        font-weight:600;font-size:.88rem;color:var(--blue);
        background:var(--surface2);border-radius:8px;
        padding:.48rem .8rem!important;
    }
    [data-testid="stExpander"]{
        border:1px solid var(--border)!important;
        border-radius:var(--r)!important;
    }

    [data-testid="metric-container"]{
        background:var(--surface);
        border:1px solid var(--border);
        border-radius:var(--r-lg);
        padding:.8rem 1rem;
        box-shadow:var(--sh0);
        transition:var(--tr);
        position:relative;overflow:hidden;
    }
    [data-testid="metric-container"]::after{
        content:'';position:absolute;top:0;left:0;right:0;height:3px;
        background:linear-gradient(90deg,var(--blue-l),var(--blue-m));
        border-radius:var(--r) var(--r) 0 0;
    }
    [data-testid="metric-container"]:hover{
        box-shadow:var(--sh1);border-color:var(--blue-b);
    }
    [data-testid="stMetricValue"]{
        font-size:1.25rem!important;font-weight:700!important;
        color:var(--blue)!important;font-family:'Inter',sans-serif!important;
    }
    [data-testid="stMetricLabel"]{
        font-size:.72rem!important;font-weight:600!important;
        color:var(--muted)!important;text-transform:uppercase;
        letter-spacing:.5px;
    }

    .stTextInput input,.stNumberInput input{
        border-radius:var(--r)!important;
        border:1.5px solid var(--border)!important;
        font-size:.86rem!important;transition:var(--tr);
        background:var(--surface)!important;
    }
    .stTextInput input:focus,.stNumberInput input:focus{
        border-color:var(--blue-l)!important;
        box-shadow:var(--glow)!important;
    }

    [data-testid="stDataFrame"],[data-testid="stDataEditor"]{
        border-radius:var(--r-lg)!important;
        border:1px solid var(--border)!important;
        overflow:hidden;
        box-shadow:var(--sh1)!important;
    }

    hr{border:none;border-top:1px solid var(--border);margin:1rem 0;}

    .ms-log-area{
        background:#080D18;
        border:1px solid rgba(59,130,246,.15);
        border-radius:var(--r-lg);
        padding:1.1rem 1.2rem;
        font-family:'JetBrains Mono',monospace;
        font-size:.75rem;color:#CBD5E1;
        max-height:420px;overflow-y:auto;
        white-space:pre-wrap;line-height:1.6;
        box-shadow:inset 0 2px 8px rgba(0,0,0,.3);
    }
    .ms-log-area .log-ts{color:#334155;}
    .ms-log-area .log-ok{color:#22D3EE;}
    .ms-log-area .log-warn{color:#F59E0B;}
    .ms-log-area .log-err{color:#F87171;}
    .ms-log-area .log-info{color:#60A5FA;}

    .cte-stat-grid{
        display:grid;
        grid-template-columns:repeat(5,1fr);
        gap:1rem;margin:1rem 0;
    }
    .cte-stat-card{
        background:var(--surface);
        border:1px solid var(--border);
        border-radius:var(--r-lg);
        padding:1.2rem 1.4rem;
        position:relative;overflow:hidden;
        transition:var(--tr);
        box-shadow:var(--sh0);
    }
    .cte-stat-card::before{
        content:'';position:absolute;
        top:0;left:0;right:0;height:3px;
        background:linear-gradient(90deg,var(--blue-l),var(--green-l));
    }
    .cte-stat-card:hover{
        box-shadow:var(--sh2);
        transform:translateY(-2px);
    }
    .cte-stat-label{
        font-size:.68rem;font-weight:700;color:var(--muted);
        text-transform:uppercase;letter-spacing:.12em;margin-bottom:.55rem;
    }
    .cte-stat-value{
        font-family:'JetBrains Mono',monospace;
        font-size:1.6rem;font-weight:600;
        color:var(--blue-l);line-height:1;
    }
    .cte-stat-sub{font-size:.72rem;color:var(--muted2);margin-top:.35rem;}

    .progress-container {
        background: var(--surface);
        border-radius: var(--r-lg);
        padding: 1.5rem;
        margin: 1rem 0;
        border: 1px solid var(--border);
    }
    .progress-bar-bg {
        background: var(--bg);
        border-radius: 10px;
        height: 24px;
        overflow: hidden;
        position: relative;
    }
    .progress-bar-fill {
        height: 100%;
        background: linear-gradient(90deg, var(--blue-l), var(--green-l));
        transition: width 0.5s ease;
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
        color: white;
        font-weight: 600;
        font-size: 0.75rem;
    }

    @media(max-width:1024px){
        .cte-stat-grid{grid-template-columns:repeat(3,1fr);}
        .hero{padding:2rem 2rem 1.8rem;}
        .hero-title{font-size:1.8rem;}
    }
    @media(max-width:768px){
        .hero{padding:1.6rem 1.2rem 1.4rem;border-radius:var(--r-xl);}
        .hero-title{font-size:1.45rem;letter-spacing:-.4px;}
        .hero-logo{max-width:140px;}
        .hero-sub{font-size:.85rem;}
        .cte-stat-grid{grid-template-columns:1fr 1fr;}
        .stTabs [data-baseweb="tab"]{padding:.35rem .6rem;font-size:.78rem;}
        .chip{font-size:.68rem;padding:.15rem .55rem;}
        .ph-title{font-size:1.1rem;}
        .block-container{padding-left:.75rem!important;padding-right:.75rem!important;}
    }
    @media(max-width:480px){
        .hero-title{font-size:1.2rem;}
        .hero{padding:1.2rem .9rem 1rem;border-radius:var(--r-lg);}
        .cte-stat-grid{grid-template-columns:1fr;}
        .hero-sub{display:none;}
    }
    </style>""")


# ==============================================================================
# PROCESSADOR CT-e OTIMIZADO PARA MÚLTIPLOS ZIPS E 50K+ XMLs
# ==============================================================================
class CTeProcessor:
    """Processador otimizado para extração de dados de CT-e em massa com suporte a múltiplos ZIPs"""
    
    def __init__(self):
        self.processed_data = []
        self.errors = []
        self.total_processed = 0
        self.total_files_scanned = 0
        self.lock = Lock()
        self._namespace_cache = {}
        self.batch_size = 1000  # Processa em lotes para evitar sobrecarga
        
    def extract_nfe_number_from_key(self, chave_acesso: str) -> Optional[str]:
        """Extrai número da NFe da chave de acesso (posições 25-34)"""
        if not chave_acesso or len(chave_acesso) != 44:
            return None
        try:
            return chave_acesso[25:34]
        except Exception:
            return None

    def extract_peso_bruto(self, root: ET.Element) -> float:
        """Extrai peso bruto do CT-e"""
        try:
            tipos_peso = ['PESO BRUTO', 'PESO BASE DE CALCULO', 'PESO BASE CALCCULO', 'PESO']
            
            for prefix, uri in CTE_NAMESPACES.items():
                for infQ in root.findall(f'.//{{{uri}}}infQ'):
                    tpMed = infQ.find(f'{{{uri}}}tpMed')
                    qCarga = infQ.find(f'{{{uri}}}qCarga')
                    if tpMed is not None and tpMed.text and qCarga is not None and qCarga.text:
                        for tp in tipos_peso:
                            if tp in tpMed.text.upper():
                                return float(qCarga.text)
            
            for infQ in root.findall('.//infQ'):
                tpMed = infQ.find('tpMed')
                qCarga = infQ.find('qCarga')
                if tpMed is not None and tpMed.text and qCarga is not None and qCarga.text:
                    for tp in tipos_peso:
                        if tp in tpMed.text.upper():
                            return float(qCarga.text)
            return 0.0
        except Exception:
            return 0.0

    def _find_text(self, element: ET.Element, xpath: str) -> Optional[str]:
        """Busca texto em elemento com suporte a namespaces"""
        try:
            for prefix, uri in CTE_NAMESPACES.items():
                found = element.find(xpath.replace('cte:', f'{{{uri}}}'))
                if found is not None and found.text:
                    return found.text.strip()
            
            found = element.find(xpath.replace('cte:', ''))
            if found is not None and found.text:
                return found.text.strip()
            
            return None
        except Exception:
            return None

    def extract_cte_data(self, xml_content: bytes, filename: str, batch_id: str = "") -> Optional[Dict]:
        """Extrai todos os dados relevantes de um arquivo CT-e"""
        try:
            try:
                root = ET.fromstring(xml_content)
            except ET.ParseError:
                try:
                    content_str = xml_content.decode('utf-8', errors='replace')
                    root = ET.fromstring(content_str)
                except Exception:
                    return None

            nCT = self._find_text(root, './/cte:nCT')
            dhEmi = self._find_text(root, './/cte:dhEmi')
            cMunIni = self._find_text(root, './/cte:cMunIni')
            UFIni = self._find_text(root, './/cte:UFIni')
            cMunFim = self._find_text(root, './/cte:cMunFim')
            UFFim = self._find_text(root, './/cte:UFFim')
            emit_xNome = self._find_text(root, './/cte:emit/cte:xNome')
            emit_CNPJ = self._find_text(root, './/cte:emit/cte:CNPJ')
            vTPrest = self._find_text(root, './/cte:vTPrest')
            rem_xNome = self._find_text(root, './/cte:rem/cte:xNome')
            dest_xNome = self._find_text(root, './/cte:dest/cte:xNome')
            dest_CNPJ = self._find_text(root, './/cte:dest/cte:CNPJ')
            dest_CPF = self._find_text(root, './/cte:dest/cte:CPF')
            dest_xLgr = self._find_text(root, './/cte:dest/cte:enderDest/cte:xLgr')
            dest_nro = self._find_text(root, './/cte:dest/cte:enderDest/cte:nro')
            dest_xBairro = self._find_text(root, './/cte:dest/cte:enderDest/cte:xBairro')
            dest_xMun = self._find_text(root, './/cte:dest/cte:enderDest/cte:xMun')
            dest_UF = self._find_text(root, './/cte:dest/cte:enderDest/cte:UF')
            dest_CEP = self._find_text(root, './/cte:dest/cte:enderDest/cte:CEP')
            infNFe_chave = self._find_text(root, './/cte:infNFe/cte:chave')
            
            numero_nfe = self.extract_nfe_number_from_key(infNFe_chave) if infNFe_chave else None
            peso_bruto = self.extract_peso_bruto(root)

            data_formatada = None
            if dhEmi:
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S'):
                    try:
                        data_formatada = datetime.strptime(dhEmi[:10], fmt).strftime('%d/%m/%Y')
                        break
                    except Exception:
                        pass
                if not data_formatada:
                    data_formatada = dhEmi[:10]

            documento_destinatario = dest_CNPJ or dest_CPF or 'N/A'
            
            endereco = ""
            if dest_xLgr:
                endereco += dest_xLgr
                if dest_nro: endereco += f", {dest_nro}"
                if dest_xBairro: endereco += f" - {dest_xBairro}"
                if dest_xMun: endereco += f", {dest_xMun}"
                if dest_UF: endereco += f"/{dest_UF}"
                if dest_CEP: endereco += f" - CEP: {dest_CEP}"
            endereco = endereco or "N/A"

            try:
                vTPrest = float(vTPrest) if vTPrest else 0.0
            except (ValueError, TypeError):
                vTPrest = 0.0

            return {
                'Arquivo': filename,
                'Lote': batch_id,
                'nCT': nCT or 'N/A',
                'Data Emissao': data_formatada or dhEmi or 'N/A',
                'Cod Municipio Inicio': cMunIni or 'N/A',
                'UF Inicio': UFIni or 'N/A',
                'Cod Municipio Fim': cMunFim or 'N/A',
                'UF Fim': UFFim or 'N/A',
                'Emitente': emit_xNome or 'N/A',
                'CNPJ Emitente': emit_CNPJ or 'N/A',
                'Valor Prestacao': vTPrest,
                'Peso Bruto (kg)': peso_bruto,
                'Remetente': rem_xNome or 'N/A',
                'Destinatario': dest_xNome or 'N/A',
                'Documento Destinatario': documento_destinatario,
                'Endereco Destinatario': endereco,
                'Municipio Destino': dest_xMun or 'N/A',
                'UF Destino': dest_UF or 'N/A',
                'Chave NFe': infNFe_chave or 'N/A',
                'Numero NFe': numero_nfe or 'N/A',
                'Data Processamento': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            }
        except Exception as e:
            logger.error(f"Erro ao processar {filename}: {e}")
            return None

    def process_xml_file(self, file_path: Path, batch_id: str = "", log_fn=None) -> Optional[Dict]:
        """Processa um arquivo XML individual"""
        try:
            with open(file_path, 'rb') as f:
                content = f.read()
            
            try:
                content_str = content.decode('utf-8', errors='replace')
                if 'CTe' not in content_str and 'conhecimento' not in content_str.lower():
                    return None
            except Exception:
                return None
            
            data = self.extract_cte_data(content, file_path.name, batch_id)
            
            if data:
                with self.lock:
                    self.processed_data.append(data)
                    self.total_processed += 1
                return data
            else:
                with self.lock:
                    self.errors.append(file_path.name)
                return None
                
        except Exception as e:
            with self.lock:
                self.errors.append(f"{file_path.name}: {str(e)}")
            if log_fn:
                log_fn(f"❌ Erro: {file_path.name} - {e}", 'err')
            return None

    def process_files_batch(self, files: List[Path], batch_id: str = "", log_fn=None, 
                           max_workers: int = 8, progress_callback=None) -> int:
        """Processa múltiplos arquivos em paralelo com controle de progresso"""
        if not files:
            return 0
        
        processed_count = 0
        total = len(files)
        
        # Processa em lotes para evitar sobrecarga
        for i in range(0, total, self.batch_size):
            batch = files[i:i + self.batch_size]
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(self.process_xml_file, f, batch_id, log_fn): f for f in batch}
                
                for future in concurrent.futures.as_completed(futures):
                    result = future.result()
                    if result:
                        processed_count += 1
                    
                    if progress_callback:
                        progress_callback(i + len(batch), total)
            
            # Força garbage collection após cada lote
            gc.collect()
        
        return processed_count

    def process_zip_file(self, zip_path: Path, batch_id: str = "", log_fn=None, 
                        max_workers: int = 8, progress_callback=None) -> int:
        """Processa um arquivo ZIP contendo XMLs"""
        try:
            temp_dir = tempfile.mkdtemp(prefix="cte_zip_")
            xml_files = []
            
            with zipfile.ZipFile(zip_path, 'r') as zf:
                for name in zf.namelist():
                    if name.lower().endswith('.xml'):
                        content = zf.read(name)
                        temp_path = Path(temp_dir) / Path(name).name
                        with open(temp_path, 'wb') as f:
                            f.write(content)
                        xml_files.append(temp_path)
            
            if log_fn:
                log_fn(f"📦 {len(xml_files)} XMLs extraídos de {zip_path.name}", 'info')
            
            result = self.process_files_batch(xml_files, batch_id, log_fn, max_workers, progress_callback)
            
            # Limpa diretório temporário
            shutil.rmtree(temp_dir, ignore_errors=True)
            
            return result
            
        except Exception as e:
            if log_fn:
                log_fn(f"❌ Erro ao processar ZIP {zip_path.name}: {e}", 'err')
            return 0

    def process_multiple_zips(self, zip_files: List[Path], log_fn=None, 
                            max_workers: int = 8, progress_callback=None) -> Dict[str, int]:
        """Processa múltiplos arquivos ZIP em sequência"""
        results = {}
        total_zips = len(zip_files)
        
        for idx, zip_path in enumerate(zip_files):
            batch_id = f"ZIP_{idx+1}_{zip_path.stem}"
            if log_fn:
                log_fn(f"📦 Processando {zip_path.name} ({idx+1}/{total_zips})...", 'info')
            
            count = self.process_zip_file(zip_path, batch_id, log_fn, max_workers, progress_callback)
            results[zip_path.name] = count
            
            if log_fn:
                log_fn(f"✅ {zip_path.name}: {count} CT-es extraídos", 'ok')
        
        return results

    def get_dataframe(self) -> Optional[pd.DataFrame]:
        """Retorna os dados como DataFrame"""
        if not self.processed_data:
            return None
        return pd.DataFrame(self.processed_data)

    def get_summary(self) -> Dict:
        """Gera resumo estatístico dos dados processados"""
        if not self.processed_data:
            return {
                'total': 0,
                'peso_total': 0.0,
                'valor_total': 0.0,
                'emitentes_unicos': 0,
                'ufs_origem': 0,
                'ufs_destino': 0,
                'errors': len(self.errors)
            }
        
        df = pd.DataFrame(self.processed_data)
        return {
            'total': len(df),
            'peso_total': df['Peso Bruto (kg)'].sum(),
            'valor_total': df['Valor Prestacao'].sum(),
            'emitentes_unicos': df['Emitente'].nunique(),
            'ufs_origem': df['UF Inicio'].nunique(),
            'ufs_destino': df['UF Destino'].nunique(),
            'errors': len(self.errors)
        }

    def export_to_excel(self) -> Optional[bytes]:
        """Exporta os dados processados para Excel com formatação profissional"""
        if not self.processed_data:
            return None
        
        df = pd.DataFrame(self.processed_data)
        buf = io.BytesIO()
        
        try:
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Dados_CTe')
                
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                ws = writer.sheets['Dados_CTe']
                
                header_fill = PatternFill('solid', start_color='0D1B2A', end_color='0D1B2A')
                header_font = Font(bold=True, color='22D3A5', name='Inter', size=10)
                border = Border(
                    left=Side(style='thin', color='1A2A3A'),
                    right=Side(style='thin', color='1A2A3A'),
                    top=Side(style='thin', color='1A2A3A'),
                    bottom=Side(style='thin', color='1A2A3A')
                )
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 4, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value in ['Valor Prestacao', 'Peso Bruto (kg)']:
                        col_letter = get_column_letter(idx)
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                            for c in row:
                                c.number_format = '#,##0.00'
                
                ws.freeze_panes = 'A2'
                
            buf.seek(0)
            return buf.getvalue()
            
        except Exception as e:
            logger.error(f"Erro ao exportar Excel: {e}")
            return None

    def clear_data(self):
        """Limpa todos os dados processados"""
        self.processed_data = []
        self.errors = []
        self.total_processed = 0
        self.total_files_scanned = 0
        gc.collect()


# ==============================================================================
# FUNÇÃO PRINCIPAL - APP CT-e
# ==============================================================================
def cte_processor_app():
    page_header("📦", "Processador de CT-e em Massa",
                "Extraia dados de milhares de XML de Conhecimento de Transporte Eletrônico")

    # Inicializa processador se não existir na session
    if 'processor' not in st.session_state:
        st.session_state.processor = CTeProcessor()
    
    processor = st.session_state.processor

    # ==========================================================================
    # TABS
    # ==========================================================================
    tab_upload, tab_analise, tab_export = st.tabs([
        "📤  Upload & Processamento",
        "📊  Análise dos Dados",
        "📥  Exportar Dados",
    ])

    # ==========================================================================
    # TAB 1 — UPLOAD & PROCESSAMENTO
    # ==========================================================================
    with tab_upload:
        section_title("📂 Importar Arquivos CT-e")

        col1, col2 = st.columns(2, gap="large")

        with col1:
            ph("""
            <div class="uzone">
                <div class="uzone-icon">📄</div>
                <div class="uzone-title">Upload de Arquivos XML</div>
                <div class="uzone-sub">Selecione um ou mais arquivos .XML</div>
            </div>""")
            
            uploaded_files = st.file_uploader(
                "Arquivos XML",
                type=['xml'],
                accept_multiple_files=True,
                key="xml_uploader"
            )

        with col2:
            ph("""
            <div class="uzone">
                <div class="uzone-icon">📦</div>
                <div class="uzone-title">Upload de Múltiplos ZIPs</div>
                <div class="uzone-sub">Selecione um ou mais arquivos .ZIP com XMLs</div>
            </div>""")
            
            zip_files = st.file_uploader(
                "Arquivos ZIP",
                type=['zip'],
                accept_multiple_files=True,
                key="zip_uploader"
            )

        # Configurações de processamento
        with st.expander("⚙️ Configurações Avançadas", expanded=False):
            col_cfg1, col_cfg2 = st.columns(2)
            
            with col_cfg1:
                max_workers = st.number_input(
                    "Threads simultâneas",
                    min_value=1,
                    max_value=16,
                    value=8,
                    help="Mais threads = processamento mais rápido, mas usa mais CPU/memória"
                )
            
            with col_cfg2:
                st.info(
                    f"💾 Memória disponível: {psutil.virtual_memory().available // (1024**2)} MB\n"
                    f"🧠 Processador: {psutil.cpu_count()} núcleos"
                )

        # Container para logs
        log_container = st.container()
        
        # Botão de processamento
        st.divider()
        col_btn1, col_btn2, col_btn3 = st.columns([2, 1, 1])
        
        with col_btn1:
            process_btn = st.button(
                "🚀  Processar Arquivos",
                type="primary",
                use_container_width=True
            )
        
        with col_btn2:
            if st.button("🗑️  Limpar Dados", use_container_width=True):
                processor.clear_data()
                st.session_state.cte_data = []
                st.session_state.cte_df = None
                st.session_state.total_processed = 0
                st.session_state.total_errors = 0
                st.rerun()

        # Processamento
        if process_btn:
            if not uploaded_files and not zip_files:
                st.warning("⚠️ Selecione pelo menos um arquivo XML ou ZIP para processar.")
            else:
                processor.clear_data()
                
                logs = []
                log_placeholder = st.empty()
                progress_placeholder = st.empty()
                
                def log_fn(msg, level='info'):
                    logs.append({'msg': msg, 'level': level})
                    if len(logs) % 10 == 0:
                        render_logs(logs, log_placeholder)
                
                def render_logs(logs_list, placeholder):
                    html = '<div class="ms-log-area">'
                    for entry in logs_list[-50:]:
                        cls = f"log-{entry['level']}"
                        html += f'<span class="{cls}">{entry["msg"]}</span>\n'
                    html += '</div>'
                    placeholder.markdown(html, unsafe_allow_html=True)
                
                def update_progress(current, total):
                    if total > 0:
                        pct = int((current / total) * 100)
                        progress_placeholder.markdown(f"""
                        <div class="progress-container">
                            <div style="display:flex;justify-content:space-between;margin-bottom:0.5rem;">
                                <span style="font-weight:600;">Progresso</span>
                                <span style="font-weight:600;color:var(--blue-l);">{pct}%</span>
                            </div>
                            <div class="progress-bar-bg">
                                <div class="progress-bar-fill" style="width:{pct}%;">
                                    {pct}%
                                </div>
                            </div>
                            <div style="text-align:center;margin-top:0.5rem;font-size:0.85rem;color:var(--muted);">
                                {current:,} de {total:,} arquivos processados
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                
                try:
                    total_files = 0
                    temp_files = []
                    temp_zip_files = []
                    all_xml_files = []
                    
                    # Processa uploads individuais
                    if uploaded_files:
                        log_fn(f"📂 Preparando {len(uploaded_files)} arquivo(s) XML...", 'info')
                        
                        temp_dir = tempfile.mkdtemp(prefix="cte_upload_")
                        for file in uploaded_files:
                            temp_path = Path(temp_dir) / file.name
                            with open(temp_path, 'wb') as f:
                                f.write(file.getvalue())
                            temp_files.append(temp_path)
                            all_xml_files.append(temp_path)
                        
                        total_files += len(uploaded_files)
                    
                    # Processa ZIPs
                    if zip_files:
                        log_fn(f"📦 Preparando {len(zip_files)} arquivo(s) ZIP...", 'info')
                        
                        zip_temp_dir = tempfile.mkdtemp(prefix="cte_zips_")
                        for file in zip_files:
                            temp_zip_path = Path(zip_temp_dir) / file.name
                            with open(temp_zip_path, 'wb') as f:
                                f.write(file.getvalue())
                            temp_zip_files.append(temp_zip_path)
                            
                            # Extrai XMLs do ZIP
                            try:
                                with zipfile.ZipFile(temp_zip_path, 'r') as zf:
                                    for name in zf.namelist():
                                        if name.lower().endswith('.xml'):
                                            content = zf.read(name)
                                            xml_name = f"{file.stem}_{Path(name).name}"
                                            xml_path = Path(zip_temp_dir) / xml_name
                                            with open(xml_path, 'wb') as f:
                                                f.write(content)
                                            all_xml_files.append(xml_path)
                                            total_files += 1
                            except Exception as e:
                                log_fn(f"⚠️ Erro ao extrair {file.name}: {e}", 'warn')
                        
                        log_fn(f"📦 {total_files} XMLs extraídos dos ZIPs", 'info')
                    
                    # Processa todos os XMLs
                    if all_xml_files:
                        log_fn(f"🚀 Iniciando processamento de {len(all_xml_files)} XMLs...", 'info')
                        
                        # Processa em lotes
                        batch_size = 1000
                        total_batches = (len(all_xml_files) + batch_size - 1) // batch_size
                        
                        for batch_idx in range(total_batches):
                            start_idx = batch_idx * batch_size
                            end_idx = min(start_idx + batch_size, len(all_xml_files))
                            batch = all_xml_files[start_idx:end_idx]
                            
                            log_fn(f"📊 Lote {batch_idx + 1}/{total_batches} ({len(batch)} arquivos)", 'info')
                            
                            count = processor.process_files_batch(
                                batch,
                                batch_id=f"LOTE_{batch_idx+1}",
                                log_fn=log_fn,
                                max_workers=max_workers,
                                progress_callback=update_progress
                            )
                            
                            # Atualiza estado
                            st.session_state.total_processed += count
                            
                            # Força GC após cada lote
                            gc.collect()
                    
                    # Resumo final
                    summary = processor.get_summary()
                    log_fn(f"", 'info')
                    log_fn(f"══════════════════════════════════════", 'info')
                    log_fn(f"📊 RESUMO FINAL", 'info')
                    log_fn(f"✅ CT-es processados: {summary['total']:,}", 'ok')
                    log_fn(f"📁 Arquivos lidos: {len(all_xml_files):,}", 'info')
                    log_fn(f"💰 Valor total: R$ {summary['valor_total']:,.2f}", 'info')
                    log_fn(f"⚖️ Peso total: {summary['peso_total']:,.0f} kg", 'info')
                    log_fn(f"🏢 Emitentes únicos: {summary['emitentes_unicos']}", 'info')
                    if summary['errors'] > 0:
                        log_fn(f"⚠️ Erros: {summary['errors']}", 'warn')
                    log_fn(f"══════════════════════════════════════", 'info')
                    
                    # Armazena no session state
                    st.session_state.cte_data = processor.processed_data
                    st.session_state.cte_df = processor.get_dataframe()
                    st.session_state.total_processed = summary['total']
                    st.session_state.total_errors = summary['errors']
                    
                    # Limpa arquivos temporários
                    for f in temp_files + all_xml_files + temp_zip_files:
                        try:
                            if Path(f).exists():
                                os.unlink(f)
                        except:
                            pass
                    
                    show_success_animation(f"Processamento concluído! {summary['total']:,} CT-es extraídos")
                    
                except Exception as e:
                    log_fn(f"❌ ERRO: {str(e)}", 'err')
                    st.error(f"Erro durante o processamento: {str(e)}")
                    st.code(traceback.format_exc())
                
                # Renderiza logs finais
                render_logs(logs, log_placeholder)
                progress_placeholder.empty()

        # Exibe logs se existirem
        if 'logs' in locals() and logs:
            st.divider()
            section_title("📋 Log de Processamento")
            render_logs(logs, st.empty())

    # ==========================================================================
    # TAB 2 — ANÁLISE DOS DADOS
    # ==========================================================================
    with tab_analise:
        df = st.session_state.cte_df
        
        if df is not None and not df.empty:
            section_title("📊 Painel de Análise")

            # Métricas
            summary = {
                'total': len(df),
                'peso_total': df['Peso Bruto (kg)'].sum(),
                'valor_total': df['Valor Prestacao'].sum(),
                'emitentes': df['Emitente'].nunique(),
                'ufs_origem': df['UF Inicio'].nunique(),
                'ufs_destino': df['UF Destino'].nunique(),
                'peso_medio': df['Peso Bruto (kg)'].mean(),
                'valor_medio': df['Valor Prestacao'].mean(),
            }

            # Cards de métricas
            ph(f"""
            <div class="cte-stat-grid">
                <div class="cte-stat-card">
                    <div class="cte-stat-label">📋 CT-es Processados</div>
                    <div class="cte-stat-value">{summary['total']:,}</div>
                    <div class="cte-stat-sub">documentos fiscais</div>
                </div>
                <div class="cte-stat-card">
                    <div class="cte-stat-label">💰 Valor Total</div>
                    <div class="cte-stat-value">R$ {summary['valor_total']:,.2f}</div>
                    <div class="cte-stat-sub">prestação de serviço</div>
                </div>
                <div class="cte-stat-card">
                    <div class="cte-stat-label">⚖️ Peso Total</div>
                    <div class="cte-stat-value">{summary['peso_total']:,.0f}</div>
                    <div class="cte-stat-sub">quilogramas</div>
                </div>
                <div class="cte-stat-card">
                    <div class="cte-stat-label">🏢 Emitentes</div>
                    <div class="cte-stat-value">{summary['emitentes']}</div>
                    <div class="cte-stat-sub">transportadoras únicas</div>
                </div>
                <div class="cte-stat-card">
                    <div class="cte-stat-label">📍 UFs</div>
                    <div class="cte-stat-value">{summary['ufs_origem']} → {summary['ufs_destino']}</div>
                    <div class="cte-stat-sub">origem → destino</div>
                </div>
            </div>
            """)

            # Filtros
            section_title("🔎 Filtros")
            fc1, fc2, fc3 = st.columns(3)
            
            with fc1:
                uf_origem = st.multiselect(
                    "UF Origem",
                    options=sorted(df['UF Inicio'].unique()),
                    key="filtro_uf_origem"
                )
            
            with fc2:
                uf_destino = st.multiselect(
                    "UF Destino",
                    options=sorted(df['UF Destino'].unique()),
                    key="filtro_uf_destino"
                )
            
            with fc3:
                emitente = st.multiselect(
                    "Emitente",
                    options=sorted(df['Emitente'].unique()),
                    key="filtro_emitente"
                )

            # Filtro de peso
            pmin = float(df['Peso Bruto (kg)'].min())
            pmax = float(df['Peso Bruto (kg)'].max())
            if pmin < pmax:
                peso_range = st.slider(
                    "Faixa de Peso (kg)",
                    pmin, pmax, (pmin, pmax),
                    format="%.1f kg",
                    key="filtro_peso"
                )
            else:
                peso_range = (pmin, pmax)

            # Aplica filtros
            fdf = df.copy()
            if uf_origem:
                fdf = fdf[fdf['UF Inicio'].isin(uf_origem)]
            if uf_destino:
                fdf = fdf[fdf['UF Destino'].isin(uf_destino)]
            if emitente:
                fdf = fdf[fdf['Emitente'].isin(emitente)]
            fdf = fdf[(fdf['Peso Bruto (kg)'] >= peso_range[0]) & 
                      (fdf['Peso Bruto (kg)'] <= peso_range[1])]

            # Métricas filtradas
            st.divider()
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("📋 Registros", len(fdf))
            m2.metric("💰 Valor Total", f"R$ {fdf['Valor Prestacao'].sum():,.2f}")
            m3.metric("⚖️ Peso Total", f"{fdf['Peso Bruto (kg)'].sum():,.2f} kg")
            m4.metric("🏢 Emitentes", fdf['Emitente'].nunique())

            # Tabela de dados
            section_title("📋 Dados")
            
            cols_to_show = ['Lote', 'Arquivo', 'nCT', 'Data Emissao', 'Emitente', 
                           'UF Inicio', 'UF Destino', 'Peso Bruto (kg)', 'Valor Prestacao']
            
            st.dataframe(
                fdf[cols_to_show],
                use_container_width=True,
                height=400,
                column_config={
                    "Valor Prestacao": st.column_config.NumberColumn(format="R$ %.2f"),
                    "Peso Bruto (kg)": st.column_config.NumberColumn(format="%.2f kg"),
                }
            )

            # Visualizações
            section_title("📈 Análise Visual")
            
            v1, v2 = st.columns(2)
            
            with v1:
                fig1 = px.histogram(
                    fdf, 
                    x='Peso Bruto (kg)',
                    nbins=30,
                    title="Distribuição de Peso Bruto",
                    color_discrete_sequence=['#3B82F6']
                )
                fig1.update_layout(
                    margin=dict(t=40, b=20, l=10, r=10),
                    height=400
                )
                st.plotly_chart(fig1, use_container_width=True)
            
            with v2:
                fig2 = px.scatter(
                    fdf,
                    x='Peso Bruto (kg)',
                    y='Valor Prestacao',
                    color='UF Destino',
                    title="Peso vs Valor da Prestação",
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                fig2.update_layout(
                    margin=dict(t=40, b=20, l=10, r=10),
                    height=400,
                    legend=dict(orientation="h", y=-0.2)
                )
                st.plotly_chart(fig2, use_container_width=True)

            # Top emitentes
            section_title("🏢 Top Emitentes")
            
            top_emitentes = fdf.groupby('Emitente').agg({
                'Valor Prestacao': 'sum',
                'Peso Bruto (kg)': 'sum',
                'Arquivo': 'count'
            }).sort_values('Valor Prestacao', ascending=False).head(10)
            
            top_emitentes.columns = ['Valor Total (R$)', 'Peso Total (kg)', 'Qtde CT-es']
            
            st.dataframe(
                top_emitentes,
                use_container_width=True,
                column_config={
                    "Valor Total (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
                    "Peso Total (kg)": st.column_config.NumberColumn(format="%.2f kg"),
                }
            )

            # Mapa de fluxos
            section_title("🗺️ Fluxo por UF")
            
            fluxo_uf = fdf.groupby(['UF Inicio', 'UF Destino']).size().reset_index(name='Quantidade')
            
            if not fluxo_uf.empty:
                fig3 = px.bar(
                    fluxo_uf,
                    x='UF Inicio',
                    y='Quantidade',
                    color='UF Destino',
                    title="Quantidade de CT-es por UF de Origem e Destino",
                    barmode='group'
                )
                fig3.update_layout(
                    margin=dict(t=40, b=30, l=10, r=10),
                    height=400,
                    legend=dict(orientation="h", y=-0.15)
                )
                st.plotly_chart(fig3, use_container_width=True)

        else:
            empty_state(
                "📊",
                "Nenhum dado processado ainda",
                "Carregue arquivos XML na aba 'Upload & Processamento'"
            )

    # ==========================================================================
    # TAB 3 — EXPORTAR DADOS
    # ==========================================================================
    with tab_export:
        df = st.session_state.cte_df
        
        if df is not None and not df.empty:
            section_title("💾 Exportar Dados")

            col_exp1, col_exp2 = st.columns([1, 2], gap="large")
            
            with col_exp1:
                st.metric("📋 Registros", f"{len(df):,}")
                st.metric("📁 Arquivos", f"{df['Arquivo'].nunique():,}")
                st.metric("📦 Lotes", f"{df['Lote'].nunique():,}" if 'Lote' in df.columns else "N/A")
                
                formato = st.radio(
                    "Formato de exportação",
                    ["📊 Excel (.xlsx)", "📄 CSV (.csv)"],
                    key="export_formato"
                )
            
            with col_exp2:
                colunas = st.multiselect(
                    "Colunas para exportar",
                    options=df.columns.tolist(),
                    default=df.columns.tolist(),
                    key="export_colunas"
                )
            
            df_export = df[colunas] if colunas else df
            
            st.divider()

            # Botão de exportação
            if formato.startswith("📊"):
                excel_bytes = processor.export_to_excel()
                
                if excel_bytes:
                    st.download_button(
                        "⬇️  Baixar Excel",
                        data=excel_bytes,
                        file_name=f"CT-e_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
                else:
                    buf = BytesIO()
                    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
                        df_export.to_excel(writer, index=False, sheet_name='CTe')
                    buf.seek(0)
                    
                    st.download_button(
                        "⬇️  Baixar Excel",
                        data=buf,
                        file_name=f"CT-e_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
            else:
                csv = df_export.to_csv(index=False, encoding='utf-8-sig')
                
                st.download_button(
                    "⬇️  Baixar CSV",
                    data=csv,
                    file_name=f"CT-e_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    type="primary",
                    use_container_width=True
                )

            # Estatísticas adicionais
            with st.expander("📊 Estatísticas Detalhadas"):
                st.dataframe(df.describe(), use_container_width=True)

            # Prévia
            with st.expander("👁️ Prévia dos dados"):
                st.dataframe(df_export.head(20), use_container_width=True)

        else:
            empty_state(
                "📥",
                "Nenhum dado disponível para exportar",
                "Processe arquivos XML na aba 'Upload & Processamento' primeiro"
            )


# ==============================================================================
# APLICAÇÃO PRINCIPAL
# ==============================================================================
def main():
    load_css()

    ph("""
    <div class="hero">
        <div class="hero-glow-left"></div>
        <img src="https://raw.githubusercontent.com/DaniloNs-creator/final/7ea6ab2a610ef8f0c11be3c34f046e7ff2cdfc6a/haefele_logo.png"
             class="hero-logo" alt="Häfele">
        <h1 class="hero-title">Processador de CT-e em Massa</h1>
        <p class="hero-sub">Extração e análise de dados de Conhecimento de Transporte Eletrônico</p>
        <div class="hero-chips">
            <span class="chip">📦 XML</span>
            <span class="chip">📁 Múltiplos ZIPs</span>
            <span class="chip">⚡ 50.000+</span>
            <span class="chip">📊 Análise</span>
            <span class="chip">📥 Exportação</span>
        </div>
    </div>""")

    cte_processor_app()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        st.error(f"Erro inesperado: {str(e)}")
        st.code(traceback.format_exc())